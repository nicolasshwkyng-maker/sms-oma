"""
SMS OMA - Script de Análisis y Corrección del Año 2026
Genera:
  1. BBDD_SMS_OMA_Corregida.xlsx  — correcciones en rojo
  2. Analisis_Factores_Comunes.xlsx — análisis exhaustivo
  3. JSON data files para la app web
"""

import pandas as pd
import numpy as np
import json
import warnings
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

warnings.filterwarnings('ignore')

BASE_DIR = "/Users/ludwinshwkyng/Documents/BBDD_SMS_OMA"
OUT_DIR  = "/Users/ludwinshwkyng/Documents/SMS_OMA_APP/output"
DATA_DIR = "/Users/ludwinshwkyng/Documents/SMS_OMA_APP/data"

# ─── Cargar datos ───────────────────────────────────────────────────────────
print("📂 Cargando datos...")
df   = pd.read_excel(f"{BASE_DIR}/BBDD_SMS_OMA.xlsx")
ata  = pd.read_excel(f"{BASE_DIR}/ATA_100.xlsx")
mat  = pd.read_excel(f"{BASE_DIR}/Matriz_SMS.xlsx")
pel  = pd.read_excel(f"{BASE_DIR}/Peligros_SRVSOP.xlsx")

df['Fecha Evento'] = pd.to_datetime(df['Fecha Evento'])
df['Year'] = df['Fecha Evento'].dt.year

# ─── Tablas de referencia ────────────────────────────────────────────────────
# Matriz de riesgo: (Severity, Likelyhood) → (RiskInd, ALARP)
risk_matrix = {}
for _, r in mat.dropna(subset=['Severity','Likelyhood3']).iterrows():
    key = (str(r['Severity']).strip(), str(r['Likelyhood3']).strip())
    risk_matrix[key] = {'RiskInd': r['RiskUnitLevel2'], 'ALARP': r['ALARP2'], 'Days': r.get('Factor13', '')}

# ATA lookup
ata_lookup = {str(r['ATA']).strip(): str(r['ATA_Sign']).strip()
              for _, r in ata.iterrows() if pd.notna(r['ATA'])}

# ─── Correcciones 2026 (basadas en análisis experto) ─────────────────────────
# Formato: {ID: {campo: nuevo_valor, 'justificacion': 'texto'}}
CORRECTIONS = {
    # ── Tipo Reporte ──────────────────────────────────────────────────────────
    323: {'Tipo Reporte': 'Reactivo',
          '_j_Tipo Reporte': 'Evento ocurrido: aeronave esperó mientras retiraban elementos del diamante de parqueo'},
    316: {'Tipo Reporte': 'Proactivo',
          '_j_Tipo Reporte': 'Condición peligrosa identificada: ausencia de gatos hidráulicos'},
    315: {'Tipo Reporte': 'Proactivo',
          '_j_Tipo Reporte': 'Condición peligrosa identificada: ausencia de gatos hidráulicos'},
    270: {'Tipo Reporte': 'Proactivo',
          '_j_Tipo Reporte': 'Acción preventiva de carga de base de datos; causa N/A indica sin evento real'},
    266: {'Tipo Reporte': 'Reactivo',
          '_j_Tipo Reporte': 'Evento ocurrido: tarea retornada con notas de corrección por formas desactualizadas'},

    # ── Peligro Genérico ──────────────────────────────────────────────────────
    328: {'Peligro Generico': 'MDA - Mantenimiento deficiente de aeronaves',
          '_j_Peligro Generico': 'Sensor de torque intermitente detectado en vuelo → falla de mantenimiento en componente de motor'},
    323: {'Peligro Generico': 'GH - Operaciones inadecuadas en tierra',
          '_j_Peligro Generico': 'Invasión del diamante de seguridad por empresa de rampa (Menzies)'},
    316: {'Peligro Generico': 'HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento',
          '_j_Peligro Generico': 'Falta de gatos hidráulicos = deficiencia en equipamiento del hangar/taller'},
    315: {'Peligro Generico': 'HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento',
          '_j_Peligro Generico': 'Falta de gatos hidráulicos para line check ATR 42-500'},
    279: {'Peligro Generico': 'MDA - Mantenimiento deficiente de aeronaves',
          '_j_Peligro Generico': 'Diferido vencido no programado en pernocta = deficiencia en planificación de mantenimiento'},
    272: {'Peligro Generico': 'MDA - Mantenimiento deficiente de aeronaves',
          '_j_Peligro Generico': 'Cadena de eventos sin cierre del NRI del FDR por 2 meses, crisis logística y de suministros'},
    270: {'Peligro Generico': 'DOC - Documentación inapropiada',
          '_j_Peligro Generico': 'Carga de base de datos TERRAIN/OBSTACLE = gestión de documentación/datos de navegación'},
    268: {'Peligro Generico': 'MDA - Mantenimiento deficiente de aeronaves',
          '_j_Peligro Generico': 'Handset roto sin dispositivo de seguridad instalado = componente con mantenimiento deficiente'},
    266: {'Peligro Generico': 'DOC - Documentación inapropiada',
          '_j_Peligro Generico': 'Tarea retornada por uso de formas desactualizadas en la documentación de mantenimiento'},
    259: {'Peligro Generico': 'HER - Uso de herramientas inapropiadas',
          '_j_Peligro Generico': 'Embudos de servicio de aceite entregados sin extensión/acoplé requerido'},
    252: {'Peligro Generico': 'PRO - Procedimientos de mantenimiento incorrectos',
          '_j_Peligro Generico': 'Modificación no autorizada (espejo en galley) = procedimiento de mantenimiento incorrecto/no aprobado'},
    250: {'Peligro Generico': 'HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento',
          '_j_Peligro Generico': 'Ausencia de sillas ergonómicas en puesto de información técnica = deficiencia de instalaciones'},
    248: {'Peligro Generico': 'HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento',
          '_j_Peligro Generico': 'Sillas de biblioteca técnica en mal estado usadas 24/7 por todo el personal'},
    246: {'Peligro Generico': 'GH - Operaciones inadecuadas en tierra',
          '_j_Peligro Generico': 'Circuito de luces de pista inoperativo en SKUC afecta operaciones de aterrizaje nocturno'},

    # ── ATA_100 ───────────────────────────────────────────────────────────────
    328: {'ATA_100': 'ATA_77', 'ATA_100: ATA_Sign': 'Engine Indicating',
          '_j_ATA_100': 'Sensor de torque del motor → ATA 77 Engine Indicating'},
    272: {'ATA_100': 'ATA_31', 'ATA_100: ATA_Sign': 'Indicating/Recording Systems',
          '_j_ATA_100': 'FDR (Flight Data Recorder) → ATA 31 Indicating/Recording Systems'},
    270: {'ATA_100': 'ATA_34', 'ATA_100: ATA_Sign': 'Navigation',
          '_j_ATA_100': 'Terrain & Obstacle Database → ATA 34 Navigation'},
    264: {'ATA_100': 'ATA_71', 'ATA_100: ATA_Sign': 'Power Plant',
          '_j_ATA_100': 'Embudo encontrado en cubierta del motor → ATA 71 Power Plant'},
    259: {'ATA_100': 'ATA_79', 'ATA_100: ATA_Sign': 'Oil',
          '_j_ATA_100': 'Servicio de aceite en motores → ATA 79 Oil'},
    258: {'ATA_100': 'ATA_31', 'ATA_100: ATA_Sign': 'Indicating/Recording Systems',
          '_j_ATA_100': 'ESIS Power Supply (Emergency Standby Instrument) → ATA 31'},
    254: {'ATA_100': 'ATA_32', 'ATA_100: ATA_Sign': 'Landing Gear',
          '_j_ATA_100': 'EMERGENCY BRAKE PRESSURE → ATA 32 Landing Gear'},
    273: {'ATA_100': 'ATA_32', 'ATA_100: ATA_Sign': 'Landing Gear',
          '_j_ATA_100': 'Lubricación tren de nariz → ATA 32 Landing Gear'},
    241: {'ATA_100': 'ATA_24', 'ATA_100: ATA_Sign': 'Electrical Power',
          '_j_ATA_100': 'Cable GPU 115V dañado → ATA 24 Electrical Power'},
    291: {'ATA_100': 'ATA_07', 'ATA_100: ATA_Sign': 'Lifting And Shoring',
          '_j_ATA_100': 'Banco de apoyo en tierra (ETAA) → ATA 07 Lifting And Shoring'},
}

# Fusionar correcciones para ID 328 (tiene tanto Peligro como ATA)
CORRECTIONS[328].update({
    'ATA_100': 'ATA_77', 'ATA_100: ATA_Sign': 'Engine Indicating',
    '_j_ATA_100': 'Sensor de torque del motor → ATA 77 Engine Indicating'
})

# ─── Mapping SPI por Peligro ──────────────────────────────────────────────────
SPI_MAP = {
    'DOC - Documentación inapropiada': {
        'default': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
        'keywords_alt': ['desactualiz', 'manual', 'texto', 'consulta'],
        'alt_spi': 'Consulta de textos desactualizados'
    },
    'PRO - Procedimientos de mantenimiento incorrectos': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'MDA - Mantenimiento deficiente de aeronaves': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'DGA - Daños graves causados a una aeronave durante las actividades de mantenimiento': 'Daños graves causados a una aeronave durante las actividades de mantenimiento',
    'GAR - Requerimientos de garantía por servicios deficientes': 'Reclamación por garantías',
    'ALM - Almacenamiento inadecuado de componentes': 'Incorrecto almacenaje de componentes aeronáuticos',
    'FH - Factores humanos como; aspectos psicosociales e incapacidad evidente del personal técnico': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'GH - Operaciones inadecuadas en tierra': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'HER - Uso de herramientas inapropiadas': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'FOD - Foreing Object Debris/Damage': 'Datos o procedimientos de mantenimiento incorrectos o deficientes',
}

def get_spi(peligro, descripcion):
    if pd.isna(peligro) or peligro == 'N/A':
        return None
    p = str(peligro).strip()
    if p not in SPI_MAP:
        return None
    spi_val = SPI_MAP[p]
    if isinstance(spi_val, dict):
        # Check if description contains keywords for alternate SPI
        desc_lower = str(descripcion).lower()
        if any(k in desc_lower for k in spi_val['keywords_alt']):
            return spi_val['alt_spi']
        return spi_val['default']
    return spi_val

# ─── Aplicar correcciones al DataFrame ───────────────────────────────────────
print("🔧 Aplicando correcciones...")
df_corrected = df.copy()

# Track changes for the changelog sheet
changes_log = []

df26_mask = df_corrected['Year'] == 2026

for idx, row in df_corrected[df26_mask].iterrows():
    rid = int(row['ID'])
    corrections_for_id = CORRECTIONS.get(rid, {})

    for field, new_val in corrections_for_id.items():
        if field.startswith('_j_'):  # skip justification keys
            continue
        old_val = row[field]
        if pd.isna(old_val) or str(old_val).strip() != str(new_val).strip():
            just_key = f'_j_{field}'
            justification = corrections_for_id.get(just_key, 'Corrección basada en análisis histórico y descripción del evento')
            changes_log.append({
                'ID': rid,
                'Fecha Evento': row['Fecha Evento'].strftime('%Y-%m-%d') if pd.notna(row['Fecha Evento']) else '',
                'Campo': field,
                'Valor Original': str(old_val) if pd.notna(old_val) else '(vacío)',
                'Valor Corregido': str(new_val),
                'Justificación': justification,
                'Área Generadora': str(row.get('Area_Generadora', '')),
                'Flota': str(row.get('Flota', '')),
            })
            df_corrected.at[idx, field] = new_val

    # Auto-completar SPI si está vacío y se conoce el Peligro
    peligro_actual = df_corrected.at[idx, 'Peligro Generico']
    spi_actual = df_corrected.at[idx, 'Indicadores SPI']
    if pd.notna(peligro_actual) and (pd.isna(spi_actual) or str(spi_actual).strip() == ''):
        new_spi = get_spi(peligro_actual, row.get('Descripción del Evento', ''))
        if new_spi:
            changes_log.append({
                'ID': rid,
                'Fecha Evento': row['Fecha Evento'].strftime('%Y-%m-%d') if pd.notna(row['Fecha Evento']) else '',
                'Campo': 'Indicadores SPI',
                'Valor Original': '(vacío)',
                'Valor Corregido': new_spi,
                'Justificación': f'SPI derivado del Peligro Genérico: {peligro_actual}',
                'Área Generadora': str(row.get('Area_Generadora', '')),
                'Flota': str(row.get('Flota', '')),
            })
            df_corrected.at[idx, 'Indicadores SPI'] = new_spi

    # Auto-completar Conciencia del Error = Pendiente si vacío y Tipo Reporte = Reactivo
    conciencia = df_corrected.at[idx, 'Conciencia del Error']
    tipo = df_corrected.at[idx, 'Tipo Reporte']
    if (pd.isna(conciencia) or str(conciencia).strip() == '') and str(tipo) == 'Reactivo':
        changes_log.append({
            'ID': rid,
            'Fecha Evento': row['Fecha Evento'].strftime('%Y-%m-%d') if pd.notna(row['Fecha Evento']) else '',
            'Campo': 'Conciencia del Error',
            'Valor Original': '(vacío)',
            'Valor Corregido': 'Pendiente',
            'Justificación': 'Registros reactivos recientes sin notificación enviada aún',
            'Área Generadora': str(row.get('Area_Generadora', '')),
            'Flota': str(row.get('Flota', '')),
        })
        df_corrected.at[idx, 'Conciencia del Error'] = 'Pendiente'

print(f"   → {len(changes_log)} correcciones registradas en {len(set(c['ID'] for c in changes_log))} registros")

# ─── Exportar Excel Corregido con celdas en ROJO ─────────────────────────────
print("📊 Generando Excel con correcciones en rojo...")

# Guardar archivo corregido primero
excel_out = f"{OUT_DIR}/BBDD_SMS_OMA_Corregida.xlsx"
# Remove cols Year
df_export = df_corrected.drop(columns=['Year'])
df_export.to_excel(excel_out, index=False, sheet_name='BBDD Corregida 2026')

# Reabrir con openpyxl para formato
wb = load_workbook(excel_out)
ws = wb.active

# Estilos
RED_FONT = Font(color='CC0000', bold=True)
RED_FILL = PatternFill('solid', fgColor='FFE8E8')
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
NORMAL_FILL = PatternFill('solid', fgColor='FFFFFF')
ALT_FILL   = PatternFill('solid', fgColor='F2F7FC')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Cabecera
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[1].height = 35

# Columnas para indexar
col_names = [cell.value for cell in ws[1]]
col_map = {name: idx+1 for idx, name in enumerate(col_names)}

# Identificar celdas corregidas
corrected_cells = set()
for c in changes_log:
    rid = c['ID']
    field = c['Campo']
    corrected_cells.add((rid, field))

# Aplicar formato a filas de datos
for row in ws.iter_rows(min_row=2):
    row_id_cell = row[col_map.get('ID', 1) - 1]
    try:
        row_id = int(row_id_cell.value) if row_id_cell.value is not None else 0
    except:
        row_id = 0

    # Determinar año de la fila
    fecha_cell = row[col_map.get('Fecha Evento', 4) - 1]
    try:
        year = pd.to_datetime(fecha_cell.value).year if fecha_cell.value else 0
    except:
        year = 0

    is_2026 = (year == 2026)
    row_fill = ALT_FILL if ws.row_dimensions[row[0].row].height else NORMAL_FILL

    for cell in row:
        col_name = col_names[cell.column - 1] if cell.column <= len(col_names) else ''
        cell.border = BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=True)

        if is_2026 and (row_id, col_name) in corrected_cells:
            cell.fill = RED_FILL
            cell.font = RED_FONT
        elif is_2026:
            cell.fill = YELLOW_FILL if cell.row % 2 == 0 else PatternFill('solid', fgColor='FFFEF7')
        else:
            cell.fill = ALT_FILL if cell.row % 2 == 0 else NORMAL_FILL

# Anchos de columna
col_widths = {
    'ID': 6, 'Estatus': 12, 'Tipo_Elemento': 12, 'Fecha Evento': 13, 'Hora del Evento': 8,
    'Origen_Reporte': 14, 'Area_Generadora': 18, 'Flota': 12, 'Matricula': 10, 'Base': 10,
    'Descripción del Evento': 50, 'Causa Probable': 40, 'Area_Causante': 16,
    'Sub_Area_Causante': 16, 'Tipo Reporte': 12, 'Peligro Generico': 35,
    'ATA_100': 10, 'ATA_100: ATA_Sign': 22, 'Descriptor': 25,
    'Consecuencias': 30, 'Defensas Actuales Para Controlar el Riesgo': 35,
    'Likelyhood': 14, 'Severity C': 14, 'RiskInd': 10, 'RiskInd: Likelyhood2': 12,
    'RiskInd: ALARP': 14, 'RiskInd: Days': 10, 'Resp. Gestion': 25,
    'Conciencia del Error': 16, 'Indicadores SPI': 45, 'Plan de Acción': 35,
    'ResRiskInd': 10, 'ResRiskInd: ALARP2': 14, 'Tolerabilidad Residual': 18,
    'Retroalimentación': 30, 'Tipo de elemento': 14, 'Ruta de acceso': 20,
}
for idx, col_name in enumerate(col_names):
    ws.column_dimensions[get_column_letter(idx+1)].width = col_widths.get(col_name, 15)

# Freeze headers
ws.freeze_panes = 'A2'

# ─── Hoja: Bitácora de Cambios ────────────────────────────────────────────────
ws_log = wb.create_sheet('Bitácora de Cambios')
log_headers = ['ID', 'Fecha Evento', 'Flota', 'Área Generadora', 'Campo Corregido', 'Valor Original', 'Valor Corregido', 'Justificación']
ws_log.append(log_headers)

for cell in ws_log[1]:
    cell.fill = PatternFill('solid', fgColor='1F4E79')
    cell.font = Font(color='FFFFFF', bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
ws_log.row_dimensions[1].height = 30

for i, c in enumerate(sorted(changes_log, key=lambda x: x['ID']), start=2):
    ws_log.append([
        c['ID'], c['Fecha Evento'], c['Flota'], c['Área Generadora'],
        c['Campo'], c['Valor Original'], c['Valor Corregido'], c['Justificación']
    ])
    row_fill = PatternFill('solid', fgColor='F8F8F8') if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for cell in ws_log[i]:
        cell.fill = row_fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    # Highlight corrected value in red
    ws_log[i][6].font = Font(color='CC0000', bold=True)

col_w_log = [6, 13, 14, 20, 22, 35, 45, 70]
for idx, w in enumerate(col_w_log, start=1):
    ws_log.column_dimensions[get_column_letter(idx)].width = w
ws_log.freeze_panes = 'A2'

wb.save(excel_out)
print(f"   ✅ Excel corregido: {excel_out}")

# ─── Análisis de Factores Comunes ─────────────────────────────────────────────
print("📈 Generando análisis de factores comunes...")

df_hist = df_corrected[df_corrected['Year'] < 2026].copy()
df26    = df_corrected[df_corrected['Year'] == 2026].copy()

analysis_out = f"{OUT_DIR}/Analisis_Factores_Comunes.xlsx"
wb2 = load_workbook(excel_out)  # start fresh

# Quitar hojas previas y crear nuevas
for sn in list(wb2.sheetnames):
    del wb2[sn]

def styled_header(ws, headers, color='1F4E79'):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

def add_data_rows(ws, data_rows, alt_color='F0F6FF'):
    for i, row_data in enumerate(data_rows, start=2):
        ws.append(row_data)
        fill = PatternFill('solid', fgColor=alt_color) if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for cell in ws[i]:
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)

# ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────────────
ws1 = wb2.create_sheet('Resumen Ejecutivo')
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 25

title_cell = ws1['A1']
title_cell.value = 'ANÁLISIS EXHAUSTIVO SMS OMA — FACTORES COMUNES 2023-2026'
title_cell.font = Font(bold=True, size=14, color='1F4E79')
title_cell.fill = PatternFill('solid', fgColor='DCE6F1')
ws1.merge_cells('A1:E1')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1['A3'].value = 'INDICADORES GLOBALES'
ws1['A3'].font = Font(bold=True, color='FFFFFF')
ws1['A3'].fill = PatternFill('solid', fgColor='2E75B6')
ws1.merge_cells('A3:E3')

kpis = [
    ('Total de Reportes (histórico)', len(df_corrected[df_corrected['Year'] < 2026]),
     '2023-2025', '', ''),
    ('Total de Reportes 2026', len(df26), 'Año en curso', '', ''),
    ('Peligro más frecuente (histórico)', df_hist['Peligro Generico'].mode()[0] if len(df_hist) > 0 else '', '', '', ''),
    ('Peligro más frecuente 2026', df26['Peligro Generico'].mode()[0] if len(df26) > 0 else '', '', '', ''),
    ('Registros con Riesgo Alto/Extremo 2026', len(df26[df26['RiskInd: ALARP'].isin(['Riesgo Alto', 'Riesgo Extremo'])]), '', '', ''),
    ('Área con más reportes 2026', df26['Area_Generadora'].mode()[0] if len(df26) > 0 else '', '', '', ''),
    ('Correcciones aplicadas 2026', len(changes_log), 'campos corregidos', '', ''),
]
for i, kpi in enumerate(kpis, start=4):
    ws1[f'A{i}'].value = kpi[0]
    ws1[f'B{i}'].value = kpi[1]
    ws1[f'C{i}'].value = kpi[2]
    for cell in [ws1[f'A{i}'], ws1[f'B{i}'], ws1[f'C{i}']]:
        cell.border = BORDER
        cell.fill = PatternFill('solid', fgColor='F2F7FF' if i % 2 == 0 else 'FFFFFF')
    ws1[f'A{i}'].font = Font(bold=True)
    ws1[f'B{i}'].font = Font(color='1F4E79', bold=True)

# ── Hoja 2: Distribución por Peligro ─────────────────────────────────────────
ws2 = wb2.create_sheet('Distribución por Peligro')
styled_header(ws2, ['Peligro Genérico', '2023', '2024', '2025', '2026 (Correg.)', 'Total', '% del Total'])
ws2.column_dimensions['A'].width = 55
for c in 'BCDEFG': ws2.column_dimensions[c].width = 14

all_peligros = sorted([p for p in df_corrected['Peligro Generico'].dropna().unique() if p != 'N/A'])
rows_p = []
for p in all_peligros:
    counts = [len(df_corrected[(df_corrected['Peligro Generico'] == p) & (df_corrected['Year'] == y)])
              for y in [2023, 2024, 2025, 2026]]
    total = sum(counts)
    pct = round(total / len(df_corrected[df_corrected['Peligro Generico'].notna()]) * 100, 1)
    rows_p.append([p] + counts + [total, f'{pct}%'])
rows_p.sort(key=lambda x: -x[5])
add_data_rows(ws2, rows_p)
# Highlight top 3
for i in range(2, min(5, len(rows_p)+2)):
    for cell in ws2[i]:
        cell.font = Font(bold=True)

# ── Hoja 3: Análisis por Área ─────────────────────────────────────────────────
ws3 = wb2.create_sheet('Análisis por Área')
styled_header(ws3, ['Área', 'Reportes 2026', 'Peligro Predominante', 'Riesgo Promedio', 'Acciones Recomendadas'])
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 45
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 60

area_recs = []
for area in df26['Area_Generadora'].dropna().unique():
    sub = df26[df26['Area_Generadora'] == area]
    count = len(sub)
    peligro_mode = sub['Peligro Generico'].mode()[0] if sub['Peligro Generico'].notna().any() else 'N/D'
    risk_vals = sub['RiskInd'].dropna()
    avg_risk = f"{risk_vals.mean():.0f}" if len(risk_vals) > 0 else 'N/D'

    # Recomendaciones basadas en el peligro predominante
    recs = {
        'DOC - Documentación inapropiada': 'Implementar revisión periódica de documentos; capacitación en uso de formas actualizadas; auditoría documental trimestral',
        'PRO - Procedimientos de mantenimiento incorrectos': 'Refuerzo en verificación de procedimientos antes de inicio de tarea; implementar doble chequeo supervisión',
        'GH - Operaciones inadecuadas en tierra': 'Coordinar con empresas de rampa; señalización del diamante de seguridad; capacitación personal de rampa',
        'MDA - Mantenimiento deficiente de aeronaves': 'Revisión del programa de mantenimiento; mejora en gestión de diferidos; seguimiento de NRIs abiertos',
        'HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento': 'Inventario y reposición de equipamiento de hangar; reporte semanal de deficiencias de instalaciones',
        'FH - Factores humanos como; aspectos psicosociales e incapacidad evidente del personal técnico': 'Programa de gestión de fatiga; evaluación psicosocial; mejora en comunicación entre turnos',
        'FOD - Foreing Object Debris/Damage': 'Inspecciones FOD programadas; campaña de concientización; canecas en puntos estratégicos del hangar',
        'HER - Uso de herramientas inapropiadas': 'Inventario y control de herramientas; solicitud de reposición de herramientas faltantes o dañadas',
    }
    rec = recs.get(peligro_mode, 'Revisar procedimientos del área y fortalecer supervisión')
    area_recs.append([area, count, peligro_mode, avg_risk, rec])

area_recs.sort(key=lambda x: -x[1])
add_data_rows(ws3, area_recs)

# ── Hoja 4: Factores Comunes y Acciones Grupales ─────────────────────────────
ws4 = wb2.create_sheet('Factores Comunes')
styled_header(ws4, ['Factor Común', 'N° Reportes Afectados', 'Áreas Involucradas', 'Nivel de Riesgo Dominante', 'Plan de Acción Compartido', 'Responsable Principal'])
for c, w in zip('ABCDEF', [35, 18, 30, 20, 70, 25]):
    ws4.column_dimensions[c].width = w

# Identificar factores comunes
factores = []

# Factor 1: Documentación desactualizada
doc_recs = df26[df26['Peligro Generico'].str.contains('DOC', na=False)]
if len(doc_recs) > 0:
    factores.append([
        'Documentación inapropiada / Formas desactualizadas',
        len(doc_recs),
        '; '.join(doc_recs['Area_Generadora'].dropna().unique()[:4]),
        doc_recs['RiskInd: ALARP'].mode()[0] if doc_recs['RiskInd: ALARP'].notna().any() else 'N/D',
        '1) Auditoría documental mensual en todas las áreas. 2) Sistema centralizado de control de versiones. 3) Alerta automática de documentos próximos a vencer. 4) Capacitación en uso correcto de formas.',
        'Biblioteca y Registros Técnicos / Ingeniería'
    ])

# Factor 2: Equipamiento de hangar insuficiente
han_recs = df26[df26['Peligro Generico'].str.contains('HAN', na=False)]
if len(han_recs) > 0:
    factores.append([
        'Deficiencias en instalaciones y equipamiento del hangar',
        len(han_recs),
        '; '.join(han_recs['Area_Generadora'].dropna().unique()[:4]),
        han_recs['RiskInd: ALARP'].mode()[0] if han_recs['RiskInd: ALARP'].notna().any() else 'N/D',
        '1) Inventario crítico de equipamiento de soporte (gatos, GPU, bancos ETAA). 2) Plan de reposición trimestral. 3) Registro de deficiencias de infraestructura con seguimiento semanal.',
        'Mantenimiento / Planeación y Programación'
    ])

# Factor 3: Operaciones inadecuadas en tierra
gh_recs = df26[df26['Peligro Generico'].str.contains('GH', na=False)]
if len(gh_recs) > 0:
    factores.append([
        'Operaciones inadecuadas en tierra (incluyendo terceros)',
        len(gh_recs),
        '; '.join(gh_recs['Area_Generadora'].dropna().unique()[:4]),
        gh_recs['RiskInd: ALARP'].mode()[0] if gh_recs['RiskInd: ALARP'].notna().any() else 'N/D',
        '1) Comunicación formal a empresas de rampa (Menzies y similares) sobre procedimientos de diamante. 2) Coordinación con aeropuertos sobre FOD en pistas. 3) Inspección pre-vuelo reforzada. 4) Reporte de incidentes de rampa a UAEAC.',
        'Seguridad Operacional / Inspección y Calidad'
    ])

# Factor 4: Mantenimiento deficiente / NRIs abiertos
mda_recs = df26[df26['Peligro Generico'].str.contains('MDA', na=False)]
if len(mda_recs) > 0:
    factores.append([
        'Mantenimiento deficiente / NRIs y diferidos sin cierre oportuno',
        len(mda_recs),
        '; '.join(mda_recs['Area_Generadora'].dropna().unique()[:4]),
        mda_recs['RiskInd: ALARP'].mode()[0] if mda_recs['RiskInd: ALARP'].notna().any() else 'N/D',
        '1) Seguimiento diario de NRIs críticos en comité. 2) Gestión de cadena de suministro para partes críticas. 3) Priorización de diferidos con vencimiento próximo. 4) Auditoría de tasa de cierre de reportes por área.',
        'Centro de Control Mantenimiento / Planeación'
    ])

# Factor 5: Procedimientos incorrectos
pro_recs = df26[df26['Peligro Generico'].str.contains('PRO', na=False)]
if len(pro_recs) > 0:
    factores.append([
        'Procedimientos de mantenimiento incorrectos',
        len(pro_recs),
        '; '.join(pro_recs['Area_Generadora'].dropna().unique()[:4]),
        pro_recs['RiskInd: ALARP'].mode()[0] if pro_recs['RiskInd: ALARP'].notna().any() else 'N/D',
        '1) Verificación obligatoria de procedimiento antes de inicio de tarea. 2) Check-list de apertura de tarea. 3) Capacitación en lectura de manuales actualizados. 4) Supervisión reforzada en tareas complejas.',
        'Inspección y Calidad / Ingeniería'
    ])

# Factor 6: FOD
fod_recs = df26[df26['Peligro Generico'].str.contains('FOD', na=False)]
if len(fod_recs) > 0:
    factores.append([
        'FOD - Objetos extraños / daños en áreas operacionales',
        len(fod_recs),
        '; '.join(fod_recs['Area_Generadora'].dropna().unique()[:4]),
        fod_recs['RiskInd: ALARP'].mode()[0] if fod_recs['RiskInd: ALARP'].notna().any() else 'N/D',
        '1) Programa FOD Walk diario en hangar y áreas de rampa. 2) Campaña de concientización mensual. 3) Instalar canecas de residuos en puntos estratégicos. 4) Protocolo de inspección post-mantenimiento.',
        'Seguridad Operacional / Mantenimiento'
    ])

factores.sort(key=lambda x: -x[1])
add_data_rows(ws4, factores)

# ── Hoja 5: Tendencias y KPIs SMS ────────────────────────────────────────────
ws5 = wb2.create_sheet('Tendencias KPIs SMS')
styled_header(ws5, ['Indicador SPI', '2023', '2024', '2025', '2026', 'Tendencia', 'Estado'])
for c, w in zip('ABCDEFG', [50, 8, 8, 8, 8, 18, 14]):
    ws5.column_dimensions[c].width = w

all_spis = [
    'Datos o procedimientos de mantenimiento incorrectos o deficientes',
    'Daños graves causados a una aeronave durante las actividades de mantenimiento',
    'Reclamación por garantías',
    'Consulta de textos desactualizados',
    'Incorrecto almacenaje de componentes aeronáuticos',
]
for spi in all_spis:
    row = [spi]
    vals = []
    for y in [2023, 2024, 2025, 2026]:
        n = len(df_corrected[(df_corrected['Indicadores SPI'] == spi) & (df_corrected['Year'] == y)])
        row.append(n)
        vals.append(n)
    # Tendencia
    if vals[-1] > vals[-2]:
        trend = '↑ Aumentando'
        state = '⚠ Atención'
    elif vals[-1] < vals[-2]:
        trend = '↓ Disminuyendo'
        state = '✓ Mejorando'
    else:
        trend = '→ Estable'
        state = '◉ Monitorear'
    row += [trend, state]
    ws5.append(row)
    i = ws5.max_row
    fill = PatternFill('solid', fgColor='F0F6FF') if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for cell in ws5[i]:
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    if '↑' in str(ws5[i][5].value):
        ws5[i][5].font = Font(color='CC0000', bold=True)
        ws5[i][6].font = Font(color='CC0000', bold=True)
    elif '↓' in str(ws5[i][5].value):
        ws5[i][5].font = Font(color='217346', bold=True)
        ws5[i][6].font = Font(color='217346', bold=True)

wb2.save(analysis_out)
print(f"   ✅ Análisis exportado: {analysis_out}")

# ─── Generar JSON para la app web ─────────────────────────────────────────────
print("🌐 Generando archivos JSON para la app web...")

# Full BBDD
records = []
for _, r in df_corrected.iterrows():
    rec = {}
    for col in df_corrected.columns:
        val = r[col]
        if pd.isna(val):
            rec[col] = None
        elif hasattr(val, 'isoformat'):
            rec[col] = val.isoformat()
        elif isinstance(val, (np.integer,)):
            rec[col] = int(val)
        elif isinstance(val, (np.floating,)):
            rec[col] = float(val) if not np.isnan(val) else None
        else:
            rec[col] = str(val).strip()
    records.append(rec)

with open(f"{DATA_DIR}/bbdd_full.json", 'w', encoding='utf-8') as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

# Corrections 2026
with open(f"{DATA_DIR}/corrections_2026.json", 'w', encoding='utf-8') as f:
    json.dump(changes_log, f, ensure_ascii=False, indent=2)

# ATA Lookup
ata_dict = []
for _, r in ata.iterrows():
    if pd.notna(r.get('ATA')):
        ata_dict.append({'code': str(r['ATA']).strip(), 'sign': str(r['ATA_Sign']).strip() if pd.notna(r['ATA_Sign']) else ''})
with open(f"{DATA_DIR}/ata_lookup.json", 'w', encoding='utf-8') as f:
    json.dump(ata_dict, f, ensure_ascii=False, indent=2)

# Peligros lookup
peligros_dict = []
for _, r in pel.iterrows():
    if pd.notna(r.get('Codigo_Peligro')):
        peligros_dict.append({
            'code': str(r['Codigo_Peligro']).strip(),
            'category': str(r.get('Categoria_Peligro', '')).strip(),
            'abbrev': str(r.get('Abreb_Peligro', '')).strip(),
            'subcat': str(r.get('Sub_Cat_Peligro', '')).strip(),
            'example': str(r.get('Ej_Peligro', '')).strip()
        })
with open(f"{DATA_DIR}/peligros_lookup.json", 'w', encoding='utf-8') as f:
    json.dump(peligros_dict, f, ensure_ascii=False, indent=2)

# Risk Matrix
risk_dict = []
for _, r in mat.iterrows():
    if pd.notna(r.get('RiskUnitLevel2')):
        risk_dict.append({
            'riskInd': float(r['RiskUnitLevel2']),
            'severity': str(r.get('Severity', '')).strip(),
            'likelyhood': str(r.get('Likelyhood3', '')).strip(),
            'alarp': str(r.get('ALARP2', '')).strip(),
            'factor12': float(r.get('Factor12', 0)),
            'factor13': float(r.get('Factor13', 0)),
        })
with open(f"{DATA_DIR}/matriz_sms.json", 'w', encoding='utf-8') as f:
    json.dump(risk_dict, f, ensure_ascii=False, indent=2)

# Training data for ML classifier (pre-2026)
training_data = []
df_train = df_corrected[df_corrected['Year'] < 2026].copy()
for _, r in df_train.iterrows():
    if pd.notna(r.get('Descripción del Evento')) and pd.notna(r.get('Peligro Generico')):
        training_data.append({
            'texto': f"{r.get('Descripción del Evento', '')} {r.get('Causa Probable', '')}",
            'area': str(r.get('Area_Generadora', '')),
            'flota': str(r.get('Flota', '')),
            'tipo_reporte': str(r.get('Tipo Reporte', '')),
            'peligro_generico': str(r.get('Peligro Generico', '')),
            'ata': str(r.get('ATA_100', '')),
            'likelyhood': str(r.get('Likelyhood', '')),
            'severity': str(r.get('Severity C', '')),
            'risk_alarp': str(r.get('RiskInd: ALARP', '')),
            'indicadores_spi': str(r.get('Indicadores SPI', '')),
            'year': int(r.get('Year', 0)),
        })
with open(f"{DATA_DIR}/training_data.json", 'w', encoding='utf-8') as f:
    json.dump(training_data, f, ensure_ascii=False, indent=2)

# Summary stats for dashboard
stats = {
    'total_records': len(df_corrected),
    'by_year': {str(y): len(df_corrected[df_corrected['Year'] == y]) for y in [2023, 2024, 2025, 2026]},
    'by_peligro': df_corrected['Peligro Generico'].value_counts().to_dict(),
    'by_area_2026': df26['Area_Generadora'].value_counts().to_dict(),
    'by_alarp_2026': df26['RiskInd: ALARP'].value_counts().to_dict(),
    'by_tipo_reporte': df_corrected['Tipo Reporte'].value_counts().to_dict(),
    'corrections_count': len(changes_log),
    'corrections_by_field': pd.Series([c['Campo'] for c in changes_log]).value_counts().to_dict(),
    'high_risk_2026': len(df26[df26['RiskInd: ALARP'].isin(['Riesgo Alto', 'Riesgo Extremo'])]),
    'peligros_options': sorted([p for p in df_corrected['Peligro Generico'].dropna().unique() if p != 'N/A']),
    'likelyhood_options': ['Frecuente', 'Probable', 'Ocasional', 'Improbable', 'Sumamente Improbable'],
    'severity_options': ['Catastrofico', 'Peligroso', 'Importante', 'Leve', 'Insignificante'],
    'spi_options': list(all_spis),
}
with open(f"{DATA_DIR}/stats.json", 'w', encoding='utf-8') as f:
    json.dump(stats, f, ensure_ascii=False, indent=2)

print(f"   ✅ JSON generados: {DATA_DIR}/")
print(f"\n🎉 PROCESO COMPLETADO")
print(f"   📄 Excel corregido:   {excel_out}")
print(f"   📊 Análisis factores: {analysis_out}")
print(f"   🌐 JSON para app:     {DATA_DIR}/")
print(f"   📝 Correcciones:      {len(changes_log)} campos en {len(set(c['ID'] for c in changes_log))} registros")
