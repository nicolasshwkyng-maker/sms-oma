"""
completar_vacios_2026.py
Rellena todos los campos vacíos de registros 2026 en la hoja 'BBDD Corregida 2026'.
Las celdas nuevas se marcan en verde. Excluye la columna 'Conciencia del Error'.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")   # verde suave (similar a Excel "Good")
GREEN_FONT = Font(color="276221")                       # texto verde oscuro

# ───────────────────────────────────────────────────────────────
#  MAPPINGS ESPECÍFICOS POR ID
# ───────────────────────────────────────────────────────────────

# Peligro Genérico - sólo 4 registros sin peligro
PELIGRO_BY_ID = {
    328: "MDA - Mantenimiento deficiente de aeronaves (instrucción no seguida, equivocación, mala práctica)",
    272: "MDA - Mantenimiento deficiente de aeronaves (instrucción no seguida, equivocación, mala práctica)",
    270: "DOC - Documentación inapropiada (manuales, procedimientos, ordenes, tarjetas de trabajo, etc.)",
    259: "HAN - Problemas con las instalaciones del hangar, taller o base auxiliar de mantenimiento",
}

# Tipo Reporte - 5 registros sin tipo
TIPO_BY_ID = {
    323: "Reactivo",
    316: "Proactivo",
    315: "Proactivo",
    270: "Proactivo",
    266: "Proactivo",
}

# ATA_100 (col 17) y ATA_Sign (col 18) — 26 registros sin ATA
ATA_BY_ID = {
    239: ("ATA_12", "Servicing"),
    243: ("ATA_05", "Time Limits/Maintenance Checks"),
    246: ("ATA_33", "Lights"),
    247: ("ATA_09", "Towing And Taxiing"),
    248: ("ATA_12", "Servicing"),
    250: ("ATA_12", "Servicing"),
    253: ("ATA_09", "Towing And Taxiing"),
    255: ("ATA_05", "Time Limits/Maintenance Checks"),
    256: ("ATA_24", "Electrical Power"),
    261: ("ATA_05", "Time Limits/Maintenance Checks"),
    266: ("ATA_05", "Time Limits/Maintenance Checks"),
    275: ("ATA_26", "Fire Protection"),
    276: ("ATA_05", "Time Limits/Maintenance Checks"),
    282: ("ATA_05", "Time Limits/Maintenance Checks"),
    290: ("ATA_05", "Time Limits/Maintenance Checks"),
    292: ("ATA_12", "Servicing"),
    313: ("ATA_12", "Servicing"),
    315: ("ATA_07", "Lifting And Shoring"),
    316: ("ATA_07", "Lifting And Shoring"),
    318: ("ATA_05", "Time Limits/Maintenance Checks"),
    320: ("ATA_09", "Towing And Taxiing"),
    323: ("ATA_09", "Towing And Taxiing"),
    324: ("ATA_12", "Servicing"),
    325: ("ATA_12", "Servicing"),
    326: ("ATA_05", "Time Limits/Maintenance Checks"),
    327: ("ATA_12", "Servicing"),
}

# Descriptor (col 19) — 11 registros
DESCRIPTOR_BY_ID = {
    251: "OMN-09",
    252: "TMA-14",
    259: "OMN-03",
    266: "ODP-02",
    268: "OMN-09",
    270: "TMD-01",
    278: "TMA-03",
    315: "OMN-03",
    316: "OMN-03",
    323: "TAO-10",
    328: "OMN-09",
}

# Area_Causante (col 13) — 24 registros
AREA_CAUSANTE_BY_ID = {
    239: "OMA",
    241: "OMA",
    242: "Ingeniería",
    243: "OMA",
    244: "OMA",
    245: "OMA",
    246: "Clientes Externos / Proveedores Externos",
    247: "Clientes Externos / Proveedores Externos",
    248: "OMA",
    249: "OMA",
    250: "OMA",
    251: "OMA",
    252: "OMA",
    253: "OMA",
    254: "OMA",
    255: "OMA",
    256: "OMA",
    257: "Ingeniería",
    258: "Ingeniería",
    259: "OMA",
    260: "Ingeniería",
    261: "Ingeniería",
    263: "OMA",
    264: "Seguridad Operacional",
}

# Sub_Area_Causante (col 14) — 32 registros
SUB_AREA_BY_ID = {
    239: "Seguridad Operacional",
    241: "Mantenimiento (TMA / Supervisor)",
    242: "Ingeniería",
    243: "Planeación y Programación",
    244: "Mantenimiento (TMA / Supervisor)",
    245: "Mantenimiento (TMA / Supervisor)",
    246: "Aeropuerto",
    247: "Aeropuerto",
    248: "Mantenimiento (TMA / Supervisor)",
    249: "Mantenimiento (TMA / Supervisor)",
    250: "Mantenimiento (TMA / Supervisor)",
    251: "Mantenimiento (TMA / Supervisor)",
    252: "Mantenimiento (TMA / Supervisor)",
    253: "Seguridad Operacional",
    254: "Seguridad Operacional",
    255: "Centro de Control Mantenimiento",
    256: "Mantenimiento (TMA / Supervisor)",
    257: "Ingeniería",
    258: "Ingeniería",
    259: "Mantenimiento (TMA / Supervisor)",
    260: "Ingeniería",
    261: "Ingeniería",
    263: "Centro de Control Mantenimiento",
    264: "Seguridad Operacional",
    # Ya tienen area_causante pero sub vacía:
    272: "Planeación y Programación",
    275: "Inspección y Calidad (Inspector)",
    276: "Biblioteca y Registros Técnicos",
    313: "Mantenimiento (TMA / Supervisor)",
    320: "Inspección y Calidad (Inspector)",
    321: "Operaciones",
    323: "Inspección y Calidad (Inspector)",
    324: "Mantenimiento (TMA / Supervisor)",
}

# Area_Generadora (col 7) — 3 registros
AREA_GEN_BY_ID = {
    321: "Operaciones",
    324: "Mantenimiento",
    325: "Mantenimiento",
}

# Matricula (col 9) — 5 registros (Flota no aplica / Todas)
MATRICULA_BY_ID = {
    313: "No Aplica",
    316: "No Aplica",
    318: "No Aplica",
    324: "No Aplica",
    325: "No Aplica",
}

# Consecuencias (col 20) — 1 registro
CONSECUENCIAS_BY_ID = {
    328: "Indicación errónea del sistema de propulsión con activación del ATPCS durante despegue",
}

# Causa Probable (col 12) — 1 registro
CAUSA_PROBABLE_BY_ID = {
    254: "Falta de seguimiento y control sobre los diferidos abiertos previo a la operación",
}

# Defensas Actuales (col 21) — 2 registros
DEFENSAS_BY_ID = {
    318: "1. (R): Procedimiento de diligenciamiento de NRI\n2. (T): Sistema de gestión de mantenimiento (ALKYM)\n3. (E): Supervisión de área por inspector de calidad",
    328: "1. (T): Sistema ATPCS como barrera automática de propulsión\n2. (R): Lista de verificación pre-vuelo de tripulación\n3. (E): Conciencia situacional y protocolo de emergencia en cabina",
}

# Indicadores SPI (col 30) — 4 registros (se completan por Peligro abajo)
SPI_BY_ID = {
    # se resuelven en la lógica por peligro; aquí sólo los que no tienen peligro asignado
    328: "Datos o procedimientos de mantenimiento incorrectos o deficientes",
}

# ───────────────────────────────────────────────────────────────
#  LÓGICA POR PELIGRO (para campos sistemáticos)
# ───────────────────────────────────────────────────────────────

SPI_BY_PELIGRO = {
    "DOC": "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "HAN": "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "GH":  "Daños graves causados a una aeronave durante las actividades de mantenimiento",
    "FOD": "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "PRO": "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "MDA": "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "FH":  "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "HER": "Datos o procedimientos de mantenimiento incorrectos o deficientes",
    "ALM": "Incorrecto almacenaje de componentes aeronáuticos",
}

PLAN_BY_PELIGRO = {
    "DOC": (
        "1. Actualizar y estandarizar la documentación técnica afectada.\n"
        "2. Capacitar al personal en el uso correcto de manuales y formularios vigentes.\n"
        "3. Implementar revisión periódica de vigencia documental mediante auditoría interna.\n"
        "4. Verificar que las impresiones de tareas correspondan a la revisión vigente."
    ),
    "HAN": (
        "1. Gestionar la adquisición o reparación de los equipos y herramientas faltantes.\n"
        "2. Señalizar y mantener despejadas las áreas de trabajo y circulación.\n"
        "3. Verificar inventario de equipos de soporte en tierra de forma semanal.\n"
        "4. Reportar necesidades de recursos al jefe de área para gestión oportuna."
    ),
    "GH": (
        "1. Coordinar con entidades externas el cumplimiento del protocolo de seguridad operacional.\n"
        "2. Demarcar y señalizar los diamantes de seguridad de las aeronaves.\n"
        "3. Socializar procedimientos de seguridad en rampa a todo el personal involucrado.\n"
        "4. Establecer punto de control para verificación previa al posicionamiento de la aeronave."
    ),
    "FOD": (
        "1. Realizar inspección FOD del área operativa inmediatamente.\n"
        "2. Concientizar al personal sobre riesgos de objetos extraños (FOD).\n"
        "3. Implementar lista de chequeo anti-FOD obligatoria previo a operaciones.\n"
        "4. Reforzar control de herramientas y materiales en áreas de trabajo."
    ),
    "MDA": (
        "1. Investigar la causa raíz del defecto de mantenimiento identificado.\n"
        "2. Revisar y verificar los procedimientos de mantenimiento aplicables al sistema.\n"
        "3. Verificar el cumplimiento de estándares de calidad en tareas similares de la flota.\n"
        "4. Implementar verificación doble (dual sign-off) para las tareas críticas identificadas."
    ),
    "PRO": (
        "1. Revisar y actualizar los procedimientos de mantenimiento afectados.\n"
        "2. Capacitar al personal técnico en los procedimientos correctos y actualizados.\n"
        "3. Implementar verificación doble obligatoria para tareas de alto impacto en aeronavegabilidad.\n"
        "4. Efectuar seguimiento al cierre de diferidos activos con atención prioritaria."
    ),
    "FH": (
        "1. Analizar los factores humanos involucrados en el evento reportado.\n"
        "2. Implementar herramientas de gestión de amenazas y errores (TEM) en las áreas afectadas.\n"
        "3. Reforzar los canales de comunicación entre áreas técnicas y de supervisión.\n"
        "4. Revisar carga de trabajo y condiciones del entorno laboral del personal involucrado."
    ),
    "ALM": (
        "1. Revisar las condiciones y procedimientos de almacenamiento de componentes.\n"
        "2. Verificar etiquetado, segregación y trazabilidad de todos los componentes en almacén.\n"
        "3. Capacitar al personal en los estándares de almacenaje aeronáutico vigentes.\n"
        "4. Auditar el área de almacén y corregir las no conformidades identificadas."
    ),
    "HER": (
        "1. Retirar del servicio la herramienta o equipo deficiente de forma inmediata.\n"
        "2. Gestionar la reparación o sustitución de la herramienta afectada.\n"
        "3. Verificar el estado general de las herramientas similares en uso.\n"
        "4. Reforzar el control y registro de estado de herramientas críticas."
    ),
}

RESP_BY_AREA = {
    "Mantenimiento":                   "Jefe de Mantenimiento",
    "Ingeniería":                      "Jefe de Ingeniería",
    "Inspección y Calidad":            "Jefe de Inspección y Calidad",
    "Seguridad Operacional":           "Jefe de Seguridad Operacional SMS",
    "Pilotos":                         "Director de Operaciones",
    "Planeación y Programación":       "Jefe de Planeación y Programación",
    "Centro de Control Mantenimiento": "Jefe Centro de Control Mantenimiento",
    "Biblioteca y Registros Técnicos": "Jefe de Registros Técnicos",
}

# Riesgo residual por ALARP actual
RESIDUAL = {
    "Riesgo Extremo": (252, "Riesgo Alto",  "MEDIO"),
    "Riesgo Alto":    (51,  "Riesgo Medio", "MINIMO"),
    "Riesgo Medio":   (25,  "Riesgo Bajo",  "MINIMO"),
    "Riesgo Bajo":    (10,  "Riesgo Bajo",  "MINIMO"),
}

DAYS_BY_ALARP = {
    "Riesgo Extremo": 2,
    "Riesgo Alto":    10,
    "Riesgo Medio":   20,
    "Riesgo Bajo":    30,
}

# ───────────────────────────────────────────────────────────────
#  FUNCIÓN AUXILIAR
# ───────────────────────────────────────────────────────────────

def is_empty(v):
    return v is None or str(v).strip() == "" or str(v).strip().lower() == "none"

def fill(ws, row, col, value, filled_cells):
    """Escribe valor sólo si la celda está vacía y registra la celda."""
    cell = ws.cell(row, col)
    if is_empty(cell.value):
        cell.value = value
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
        filled_cells.append((row, col, value))

# ───────────────────────────────────────────────────────────────
#  MAIN
# ───────────────────────────────────────────────────────────────

def main():
    path = "output/BBDD_SMS_OMA_Corregida.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["BBDD Corregida 2026"]

    # Columnas (1-indexed)
    COL = {h: i for i, h in enumerate(
        [ws.cell(1, c).value for c in range(1, ws.max_column + 1)], 1
    )}

    SKIP_COL = COL.get("Conciencia del Error", 29)   # excluir siempre

    # Filas 2026
    rows_2026 = [r for r in range(2, ws.max_row + 1)
                 if str(ws.cell(r, COL["Fecha Evento"]).value or "").startswith("2026")]

    filled_cells = []
    total_filled = 0

    for row in rows_2026:
        rid = ws.cell(row, COL["ID"]).value

        # ---------- PELIGRO GENÉRICO ----------
        if rid in PELIGRO_BY_ID:
            fill(ws, row, COL["Peligro Generico"], PELIGRO_BY_ID[rid], filled_cells)

        pg_full = str(ws.cell(row, COL["Peligro Generico"]).value or "")
        pg_code = pg_full.split(" - ")[0].strip()   # e.g. "DOC", "GH", "MDA"

        # ---------- TIPO REPORTE ----------
        if rid in TIPO_BY_ID:
            fill(ws, row, COL["Tipo Reporte"], TIPO_BY_ID[rid], filled_cells)

        # ---------- ATA ----------
        if rid in ATA_BY_ID:
            ata, sign = ATA_BY_ID[rid]
            fill(ws, row, COL["ATA_100"], ata, filled_cells)
            fill(ws, row, COL["ATA_100: ATA_Sign"], sign, filled_cells)

        # ---------- DESCRIPTOR ----------
        if rid in DESCRIPTOR_BY_ID:
            fill(ws, row, COL["Descriptor"], DESCRIPTOR_BY_ID[rid], filled_cells)

        # ---------- AREA_CAUSANTE ----------
        if rid in AREA_CAUSANTE_BY_ID:
            fill(ws, row, COL["Area_Causante"], AREA_CAUSANTE_BY_ID[rid], filled_cells)

        # ---------- SUB_AREA_CAUSANTE ----------
        if rid in SUB_AREA_BY_ID:
            fill(ws, row, COL["Sub_Area_Causante"], SUB_AREA_BY_ID[rid], filled_cells)

        # ---------- AREA_GENERADORA ----------
        if rid in AREA_GEN_BY_ID:
            fill(ws, row, COL["Area_Generadora"], AREA_GEN_BY_ID[rid], filled_cells)

        # ---------- MATRICULA ----------
        if rid in MATRICULA_BY_ID:
            fill(ws, row, COL["Matricula"], MATRICULA_BY_ID[rid], filled_cells)

        # ---------- HORA DEL EVENTO ----------
        fill(ws, row, COL["Hora del Evento"], "No registrada", filled_cells)

        # ---------- CONSECUENCIAS ----------
        if rid in CONSECUENCIAS_BY_ID:
            fill(ws, row, COL["Consecuencias"], CONSECUENCIAS_BY_ID[rid], filled_cells)

        # ---------- CAUSA PROBABLE ----------
        if rid in CAUSA_PROBABLE_BY_ID:
            fill(ws, row, COL["Causa Probable"], CAUSA_PROBABLE_BY_ID[rid], filled_cells)

        # ---------- DEFENSAS ----------
        if rid in DEFENSAS_BY_ID:
            fill(ws, row, COL["Defensas Actuales Para Controlar el Riesgo"], DEFENSAS_BY_ID[rid], filled_cells)

        # ---------- INDICADORES SPI ----------
        if rid in SPI_BY_ID:
            fill(ws, row, COL["Indicadores SPI"], SPI_BY_ID[rid], filled_cells)
        elif pg_code in SPI_BY_PELIGRO:
            fill(ws, row, COL["Indicadores SPI"], SPI_BY_PELIGRO[pg_code], filled_cells)

        # ---------- RESP. GESTION ----------
        area_gen = str(ws.cell(row, COL["Area_Generadora"]).value or "")
        resp = RESP_BY_AREA.get(area_gen, "Jefe de Seguridad Operacional SMS")
        fill(ws, row, COL["Resp. Gestion"], resp, filled_cells)

        # ---------- PLAN DE ACCIÓN ----------
        plan = PLAN_BY_PELIGRO.get(pg_code,
            "1. Investigar causa raíz del evento reportado.\n"
            "2. Implementar acciones correctivas y preventivas.\n"
            "3. Verificar efectividad de las medidas implementadas.\n"
            "4. Documentar y socializar lecciones aprendidas al personal."
        )
        fill(ws, row, COL["Plan de Acción"], plan, filled_cells)

        # ---------- RiskInd: Likelyhood2 ----------
        fill(ws, row, COL["RiskInd: Likelyhood2"], "50% efectivas", filled_cells)

        # ---------- RiskInd: Days ----------
        alarp = str(ws.cell(row, COL["RiskInd: ALARP"]).value or "")
        days = DAYS_BY_ALARP.get(alarp, 20)
        fill(ws, row, COL["RiskInd: Days"], days, filled_cells)

        # ---------- RESIDUAL RISK ----------
        res_ind, alarp2, tolerab = RESIDUAL.get(alarp, (25, "Riesgo Bajo", "MINIMO"))
        fill(ws, row, COL["ResRiskInd"], res_ind, filled_cells)
        fill(ws, row, COL["ResRiskInd: ALARP2"], alarp2, filled_cells)
        fill(ws, row, COL["Tolerabilidad Residual"], tolerab, filled_cells)

    # ---------- ESTADÍSTICAS ----------
    total_filled = len(filled_cells)
    print(f"✅ Celdas completadas: {total_filled}")

    # Contar por columna
    col_counts = {}
    for r, c, v in filled_cells:
        col_counts[c] = col_counts.get(c, 0) + 1

    headers = {i: ws.cell(1, i).value for i in range(1, ws.max_column + 1)}
    print("\nResumen por columna:")
    for c in sorted(col_counts.keys()):
        print(f"  [{c:2d}] {str(headers[c]):<45} → {col_counts[c]} celdas")

    wb.save(path)
    print(f"\n📁 Guardado: {path}")

if __name__ == "__main__":
    main()
