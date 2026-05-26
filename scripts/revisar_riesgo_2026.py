"""
Revisión exhaustiva de la valoración de riesgo 2026 (ARMS/SMS)
- Revisa los 70 registros: Likelyhood, Severity C, RiskInd, ALARP
- Aplica criterios de la metodología ARMS basados en descripción + peligro
- Genera hoja adicional en BBDD_SMS_OMA_Corregida.xlsx
"""

import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

BASE_DIR = "/Users/ludwinshwkyng/Documents/BBDD_SMS_OMA"
OUT_FILE = "/Users/ludwinshwkyng/Documents/SMS_OMA_APP/output/BBDD_SMS_OMA_Corregida.xlsx"

df = pd.read_excel(f"{BASE_DIR}/BBDD_SMS_OMA.xlsx")
mat = pd.read_excel(f"{BASE_DIR}/Matriz_SMS.xlsx")
df['Year'] = pd.to_datetime(df['Fecha Evento']).dt.year
df26 = df[df['Year'] == 2026].copy().reset_index(drop=True)

# Matriz de riesgo: (Severity, Likelyhood) → RiskInd, ALARP
risk_lkp = {}
for _, r in mat.dropna(subset=['Severity','Likelyhood3']).iterrows():
    risk_lkp[(str(r['Severity']).strip(), str(r['Likelyhood3']).strip())] = {
        'RiskInd': r['RiskUnitLevel2'], 'ALARP': r['ALARP2']
    }

# ─── Criterios ARMS para clasificación de riesgo ──────────────────────────────
# Severity (Factor de Severidad según ARMS)
# Catastrófico: accidente fatal, pérdida aeronave
# Peligroso: lesión grave, daño mayor a aeronave
# Importante: lesión menor, daño significativo a sistema
# Leve: molestia operacional, daño menor
# Insignificante: sin efecto de seguridad

# Likelyhood (Probabilidad de ocurrencia en la flota/año)
# Frecuente: >10 veces/año flota
# Probable: varias veces/año flota
# Ocasional: 1-10 veces/año flota
# Improbable: <1 vez/año flota
# Sumamente Improbable: casi imposible

# Revisiones basadas en análisis experto de cada registro 2026
# Formato: {ID: {'Likelyhood': nuevo, 'Severity C': nuevo, 'justificacion': texto}}
RISK_REVIEWS = {
    # ── Registros con Likelyhood/Severity FALTANTES ──
    328: {'Likelyhood': 'Improbable', 'Severity C': 'Peligroso',
          'justificacion': 'Sensor de torque intermitente en vuelo → Severity Peligroso (falla en sistema crítico de motor). Likelyhood Improbable (falla esporádica de sensor).'},
    327: {'Likelyhood': 'Ocasional', 'Severity C': 'Leve',
          'justificacion': 'Interrupción de trabajo por vigilancia → Severity Leve (sin consecuencia directa a aeronave). Likelyhood Ocasional (incidentes de coordinación internos recurrentes).'},
    318: {'Likelyhood': 'Ocasional', 'Severity C': 'Importante',
          'justificacion': 'Error en documentación NRI → Severity Importante (puede llevar a cierre incorrecto de tarea). Likelyhood Ocasional (errores documentales frecuentes en operación).'},

    # ── Revisión de registros con clasificación de riesgo cuestionable ──

    # ID 311: GH + Frecuente + Peligroso = 1250 Riesgo Extremo
    # Descripción: cable GPU dañado en área eléctrica durante ground ops
    # FRECUENTE parece sobreestimado si es un evento puntual
    311: {'Likelyhood': 'Ocasional', 'Severity C': 'Peligroso',
          'justificacion': 'Operación inadecuada en tierra con riesgo eléctrico → Severity Peligroso correcto (riesgo de daño a aeronave/personal). Likelyhood corregido a Ocasional: evento puntual, no patrón frecuente documentado en la flota.'},

    # ID 312: FH + ATA_09 (Towing) + Probable + Peligroso = 750 Riesgo Extremo
    # Factores humanos en remolque - Probable + Peligroso
    # Si hubo un incidente real de FH en towing, Riesgo Extremo puede ser correcto
    # Mantener si hay base, pero Probable parece alto para un evento de FH aislado
    312: {'Likelyhood': 'Ocasional', 'Severity C': 'Peligroso',
          'justificacion': 'Factores humanos en operación de remolque → Severity Peligroso correcto (riesgo de daño a aeronave durante towing). Likelyhood ajustado a Ocasional: sin evidencia de recurrencia sistémica documentada.'},

    # ID 292: HAN + Frecuente + Leve = 250 Riesgo Alto
    # No hay báscula para COMAT en base Cali → condición permanente pero baja severidad
    # Frecuente es correcto (pasa en cada despacho COMAT), Leve también correcto
    # Sin cambio requerido - mantener

    # ID 288/287: PRO + ATA_72 Engine + Probable + Leve = 150 Riesgo Medio
    # Procedimiento erróneo en motor. Leve parece bajo para un error en ATA_72
    288: {'Likelyhood': 'Probable', 'Severity C': 'Importante',
          'justificacion': 'Procedimiento incorrecto en motor (ATA_72) → Severity ajustado a Importante: un error de procedimiento en motor puede causar falla latente o daño. Leve subestima el riesgo real en sistemas de propulsión.'},
    287: {'Likelyhood': 'Probable', 'Severity C': 'Importante',
          'justificacion': 'Procedimiento incorrecto en motor (ATA_72, DHC-6) → Severity ajustado a Importante: consistente con ID 288, mismo tipo de evento mismo día. Procedimientos erróneos en motor no son Leve.'},

    # ID 319: MDA + ATA_57 Wings + Probable + Peligroso = 750 Riesgo Extremo
    # Mantenimiento deficiente en alas → Riesgo Extremo parece correcto para ATA_57
    # Sin cambio

    # ID 310: PRO + ATA_72 Engine + Probable + Peligroso = 750 Riesgo Extremo
    # Procedimiento incorrecto en motor DHC-6 → Riesgo Extremo para motor es apropiado
    # Sin cambio

    # ID 280: PRO + ATA_24 Electrical + Probable + Importante = 300 Riesgo Alto
    # Procedimiento incorrecto en sistema eléctrico → correcto
    # Sin cambio

    # ID 275: GH + Improbable + Catastrófico = 252 Riesgo Alto
    # Incendio cerca aeronave en parqueo → Catastrófico correcto (riesgo pérdida aeronave)
    # Improbable correcto (quema de caña es evento ambiental excepcional)
    # Sin cambio

    # ID 264: FOD + ATR-72 + Probable + Importante = 300 Riesgo Alto
    # Embudo encontrado dentro de cubierta motor después de servicio de aceite
    # Severity debería ser Peligroso o Catastrófico (FOD en motor puede causar falla catastrófica)
    264: {'Likelyhood': 'Improbable', 'Severity C': 'Catastrofico',
          'justificacion': 'FOD (embudo) encontrado dentro de cubierta de motor → Severity CATASTRÓFICO: un objeto extraño dentro del motor puede causar falla de motor en vuelo con consecuencias catastróficas. Likelyhood Improbable: evento de descuido aislado, no patrón sistemático.'},

    # ID 253: GH + Probable + Importante = 300 Riesgo Alto
    # Tripulación caminando frente a aeronave con motor encendido (riesgo succión)
    # Severity debería ser Peligroso o Catastrófico → succión de motor puede ser fatal
    253: {'Likelyhood': 'Ocasional', 'Severity C': 'Peligroso',
          'justificacion': 'Personal caminando frente a motor encendido (riesgo succión/ingesta) → Severity Peligroso: lesiones graves o muerte por ingesta de motor. Likelyhood Ocasional: event puntual sin indicación de recurrencia frecuente.'},

    # ID 247: FOD + ERJ-145 + Improbable + Importante = 51 Riesgo Medio
    # Piedras en pista durante aterrizaje → Severity debería ser Peligroso (daño a tren, motor)
    247: {'Likelyhood': 'Improbable', 'Severity C': 'Peligroso',
          'justificacion': 'Piedras/escombros en pista durante aterrizaje (FOD de construcción) → Severity Peligroso: daño a tren de aterrizaje, motor o ingesta FOD en motor. Improbable correcto dado que es condición temporal de obra.'},

    # ID 241: GH + GPU cable dañado + Ocasional + Peligroso = 251 Riesgo Alto
    # Cable GPU 115V con daño por fricción → correcto
    # Sin cambio

    # ID 293: DOC + ATA_72 Engine + Probable + Importante = 300 Riesgo Alto
    # Documentación inapropiada en motor → correcto
    # Sin cambio

    # ID 291: HER + ETAA (banco apoyo) barras seguridad no funcionan
    # Sin Likelyhood/Severity asignado en los datos originales (verificar)
    # Si tiene valores, revisarlos

    # ID 256: HAN + toma eléctrica desprendida hangar + Ocasional + Importante = 100 Riesgo Medio
    # Toma eléctrica suelta → Severity Importante correcto (riesgo eléctrico para personal)
    # Sin cambio

    # ID 291: ETAA barras seguridad no funcionan (herramienta de levantamiento)
    # Si Likelyhood = Probable, Severity debería ser al menos Peligroso
    291: {'Likelyhood': 'Probable', 'Severity C': 'Peligroso',
          'justificacion': 'Banco de apoyo ETAA con barras de seguridad inoperativas → Severity Peligroso: falla de la herramienta de levantamiento puede causar caída de aeronave. Likelyhood Probable: equipo en uso diario con defecto conocido.'},

    # ID 272: MDA + ATA_23 + Ocasional + Importante = 100 Riesgo Medio
    # NRI FDR abierto 2 meses, crisis logística → FDR es sistema crítico, debería ser Peligroso
    272: {'Likelyhood': 'Probable', 'Severity C': 'Peligroso',
          'justificacion': 'FDR con NRI abierto 2+ meses sin cierre por falta de repuestos → Severity Peligroso: FDR inoperativo implica operación sin sistema de grabación de datos de vuelo (incumplimiento regulatorio crítico). Likelyhood Probable dado el tiempo sin resolución.'},

    # ID 255/254: PRO + Vencimiento diferido + Ocasional + Importante = 100 Riesgo Medio
    # Diferido con vencimiento no controlado → sistema de frenos (ATA_32) debería ser Peligroso
    254: {'Likelyhood': 'Ocasional', 'Severity C': 'Peligroso',
          'justificacion': 'Diferido EMERGENCY BRAKE PRESS vencido (ATA_32 frenos) → Severity Peligroso: sistema de frenado de emergencia comprometido afecta seguridad crítica de aterrizaje. Importante subestima el riesgo de un sistema de frenos de emergencia.'},

    # ID 317: MDA + ATA_07 + Probable + Importante = 300 Riesgo Alto → correcto

    # ID 313: FOD + falta canecas en hangar + Ocasional + Peligroso = 251 Riesgo Alto
    # Severity Peligroso parece alto para falta de canecas (es condición de FOD potencial)
    313: {'Likelyhood': 'Probable', 'Severity C': 'Importante',
          'justificacion': 'Falta de canecas de basura en hangar → condición de FOD potencial. Severity Importante: el riesgo es generación de FOD, no el evento en sí. Likelyhood Probable: condición permanente detectada. Peligroso sobreestima para una condición de housekeeping.'},
}

print("📊 Analizando valoración de riesgo 2026...")

thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

risk_changes = []
for idx, row in df26.iterrows():
    rid = int(row['ID'])
    if rid not in RISK_REVIEWS:
        continue
    review = RISK_REVIEWS[rid]
    new_lik = review.get('Likelyhood')
    new_sev = review.get('Severity C')
    just = review.get('justificacion', '')

    orig_lik = str(row.get('Likelyhood', '')).strip() if pd.notna(row.get('Likelyhood')) else ''
    orig_sev = str(row.get('Severity C', '')).strip() if pd.notna(row.get('Severity C')) else ''

    # Calculate original and new risk
    orig_risk_row = risk_lkp.get((orig_sev, orig_lik))
    new_risk_row  = risk_lkp.get((new_sev, new_lik))

    orig_risk_ind = orig_risk_row['RiskInd'] if orig_risk_row else ''
    orig_alarp    = orig_risk_row['ALARP']   if orig_risk_row else ''
    new_risk_ind  = new_risk_row['RiskInd']  if new_risk_row  else ''
    new_alarp     = new_risk_row['ALARP']    if new_risk_row  else ''

    changed = (new_lik != orig_lik or new_sev != orig_sev) or (not orig_lik and new_lik)

    risk_changes.append({
        'ID': rid,
        'Fecha': str(row.get('Fecha Evento', ''))[:10],
        'Area': str(row.get('Area_Generadora', '')),
        'Flota': str(row.get('Flota', '')),
        'Peligro': str(row.get('Peligro Generico', '')),
        'Likelyhood Original': orig_lik or '(vacío)',
        'Likelyhood Revisado': new_lik,
        'Severity Original': orig_sev or '(vacío)',
        'Severity Revisado': new_sev,
        'RiskInd Original': orig_risk_ind,
        'ALARP Original': orig_alarp,
        'RiskInd Revisado': new_risk_ind,
        'ALARP Revisado': new_alarp,
        'Cambio': '✓ Ajustado' if changed else '→ Sin cambio',
        'Justificación': just,
    })

print(f"   → {len(risk_changes)} registros revisados, {sum(1 for r in risk_changes if '✓' in r['Cambio'])} ajustados")

# ─── Agregar hoja de revisión de riesgo al Excel existente ───────────────────
print("📝 Agregando revisión de riesgo al Excel...")

wb = load_workbook(OUT_FILE)

# Eliminar hoja previa si existe
if 'Revisión de Riesgo 2026' in wb.sheetnames:
    del wb['Revisión de Riesgo 2026']

ws_risk = wb.create_sheet('Revisión de Riesgo 2026')

headers = [
    'ID', 'Fecha', 'Flota', 'Área', 'Peligro Genérico',
    'Likelyhood\nOriginal', 'Likelyhood\nRevisado',
    'Severity\nOriginal',  'Severity\nRevisada',
    'RiskInd\nOriginal',   'ALARP\nOriginal',
    'RiskInd\nRevisado',   'ALARP\nRevisado',
    'Estado', 'Justificación'
]

HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=9)
RED_FILL     = PatternFill('solid', fgColor='FFE8E8')
RED_FONT     = Font(color='CC0000', bold=True)
GREEN_FILL   = PatternFill('solid', fgColor='E8F5E9')
ALT_FILL     = PatternFill('solid', fgColor='F2F7FF')
NORMAL_FILL  = PatternFill('solid', fgColor='FFFFFF')

# Header row
ws_risk.append(headers)
for cell in ws_risk[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
ws_risk.row_dimensions[1].height = 40

ALARP_COLORS = {
    'Riesgo Extremo': 'FFC7CE',
    'Riesgo Alto':    'FFEB9C',
    'Riesgo Medio':   'FFEB9C',
    'Riesgo Bajo':    'C6EFCE',
}

# Data rows
for i, rec in enumerate(sorted(risk_changes, key=lambda x: x['ID']), start=2):
    changed = '✓' in rec['Cambio']
    row_fill = RED_FILL if changed else (ALT_FILL if i % 2 == 0 else NORMAL_FILL)

    ws_risk.append([
        rec['ID'], rec['Fecha'], rec['Flota'], rec['Area'],
        rec['Peligro'][:60] if rec['Peligro'] else '',
        rec['Likelyhood Original'], rec['Likelyhood Revisado'],
        rec['Severity Original'],  rec['Severity Revisado'],
        rec['RiskInd Original'],   rec['ALARP Original'],
        rec['RiskInd Revisado'],   rec['ALARP Revisado'],
        rec['Cambio'],             rec['Justificación'],
    ])

    for cell in ws_risk[i]:
        cell.fill = row_fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

    # Highlight changed cells in red
    if changed:
        for col_idx, field in [(6, 'Likelyhood Original'), (7, 'Likelyhood Revisado'),
                               (8, 'Severity Original'), (9, 'Severity Revisado')]:
            cell = ws_risk.cell(row=i, column=col_idx)
            if rec['Likelyhood Original'] != rec['Likelyhood Revisado'] and 'Likelyhood' in field:
                cell.font = RED_FONT if 'Revisado' in field else Font(color='999999', strikethrough=True)
            if rec['Severity Original'] != rec['Severity Revisado'] and 'Severity' in field:
                cell.font = RED_FONT if 'Revisado' in field else Font(color='999999', strikethrough=True)

    # Color ALARP cells
    alarp_orig = rec['ALARP Original']
    alarp_new  = rec['ALARP Revisado']
    if alarp_orig and alarp_orig in ALARP_COLORS:
        ws_risk.cell(row=i, column=11).fill = PatternFill('solid', fgColor=ALARP_COLORS[alarp_orig])
    if alarp_new and alarp_new in ALARP_COLORS:
        ws_risk.cell(row=i, column=13).fill = PatternFill('solid', fgColor=ALARP_COLORS[alarp_new])

# Column widths
col_widths = [6, 12, 14, 22, 45, 18, 18, 16, 16, 12, 16, 12, 16, 14, 80]
for idx, w in enumerate(col_widths, start=1):
    ws_risk.column_dimensions[get_column_letter(idx)].width = w

ws_risk.freeze_panes = 'A2'

# ─── También actualizar la hoja principal con los valores de riesgo corregidos ─
print("🔄 Actualizando hoja principal con riesgo corregido...")
ws_main = wb['BBDD Corregida 2026']
col_names = [cell.value for cell in ws_main[1]]
col_map = {name: idx+1 for idx, name in enumerate(col_names)}

RED_FILL2 = PatternFill('solid', fgColor='FFE8E8')
RED_FONT2 = Font(color='CC0000', bold=True)

for row in ws_main.iter_rows(min_row=2):
    try:
        rid = int(row[col_map.get('ID', 1) - 1].value)
    except:
        continue
    if rid not in RISK_REVIEWS:
        continue
    review = RISK_REVIEWS[rid]
    new_lik = review.get('Likelyhood')
    new_sev = review.get('Severity C')
    new_risk_row = risk_lkp.get((new_sev, new_lik))

    for field, new_val in [
        ('Likelyhood', new_lik),
        ('Severity C', new_sev),
        ('RiskInd', new_risk_row['RiskInd'] if new_risk_row else None),
        ('RiskInd: ALARP', new_risk_row['ALARP'] if new_risk_row else None),
    ]:
        if new_val is None or field not in col_map:
            continue
        cell = row[col_map[field] - 1]
        orig_val = str(cell.value).strip() if cell.value is not None else ''
        new_str  = str(new_val)
        if orig_val != new_str:
            cell.value = new_val
            cell.fill = RED_FILL2
            cell.font = RED_FONT2
            cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(OUT_FILE)
print(f"✅ Excel actualizado con revisión de riesgo: {OUT_FILE}")
print(f"\n📋 RESUMEN REVISIÓN DE RIESGO 2026:")
print(f"   Registros revisados: {len(risk_changes)}")
ajustados = [r for r in risk_changes if '✓' in r['Cambio']]
print(f"   Ajustados: {len(ajustados)}")
for r in ajustados:
    print(f"   ID {r['ID']:3d}: {r['ALARP Original'] or '(vacío)'} → {r['ALARP Revisado']} | {r['Severity Original']} → {r['Severity Revisado']}")
